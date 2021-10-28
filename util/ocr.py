"""
    For video OCR of raid Battle Log to record raid attacks
"""

import csv
import io
import logging
import re
import sys
from datetime import datetime as dt

import av
import cv2
import numpy as np
import openpyxl
import pandas as pd
import pytesseract
from PIL import Image
from fuzzywuzzy import process
from openpyxl.drawing.image import Image as xlImage
from openpyxl.utils.dataframe import dataframe_to_rows
from tqdm import tqdm

logging.getLogger().setLevel(logging.ERROR)

# MEMBERS = ['ChiaSeedy', 'Petrae', 'Ivy', 'Foxen', 'Rain', 'Sweetchili',
#            'Spacelion', 'Outrageous', 'Knyne', 'Hands_UP', 'Splacto',
#            'monkey16c', 'Spanners', 'Kratos', 'Artse', 'Bixlow', 'ET',
#            'Cypress', 'r3ntuh', 'Lance', 'Happy', 'Zero', 'Eir', 'Weiss',
#            'Liz', 'DoubleCast', 'Syr', 'LordKoch', 'Zangish']

MEMBERS = ['Venus', 'Kat', 'Zap', 'Pepeg ', 'DrPigeon', 'Umbra', 'UrekMazino',
           'Bartolovic', 'Iscariots', 'jarrod', 'Bob', 'Duhboss', 'Akira',
           'Reroll', 'Hmmmm', 'Lemon', 'Myth', 'Nina', 'Percy', 'qwerty',
           'Lenka', 'Jide', 'Ai', 'Saber', 'nParagon', 'Juicetin', 'Petrae',
           'Zangish']

BOSSES = ['Cyborg Erina', 'Nine-tailed Fox Garam', 'Ancient Demon',
          'Altered Mad Panda MK-3']


def skip(iterable, num):
    try:
        while True:
            for i in range(num - 1):
                next(iterable)
            yield next(iterable)
    except StopIteration:
        return


def clean_dmg(row):
    dmg_match = re.match(r'^[\d, ]+$', row)
    if not dmg_match:
        return
    else:
        dmg_str = dmg_match.group(0)
    cleaned_dmg_str = dmg_str.replace(',', '').replace(' ', '')
    try:
        return int(cleaned_dmg_str)
    except ValueError:
        return None


def extract_name(output):
    rows = output.split('\n')
    for row in rows:
        res = process.extractOne(row, MEMBERS)
        if res[1] > 80:
            return res[0]


def extract_boss(output):
    rows = output.split('\n')
    for row in rows:
        lvl, *boss_name = row.split(' ')
        boss_name = ' '.join(boss_name)
        res = process.extractOne(boss_name, BOSSES)
        if res[1] > 90:
            return res[0], lvl


with av.open(sys.argv[1]) as container:
    stream = container.streams.video[0]

    error = 1
    appeared = set()

    df = pd.DataFrame(columns=["name", "damage", "boss", "level"])
    dmg_left = None
    dmg_prev = 0

    N = 5
    iters = stream.frames // N  # not exact
    for frame in tqdm(skip(container.decode(stream), N), total=iters):
        if not isinstance(frame, av.video.frame.VideoFrame):
            continue

        img = frame.to_image()
        img = img.transpose(Image.ROTATE_270 if len(sys.argv) > 2 else Image.ROTATE_90)

        w, h = img.size
        img = img.crop((420, 120, w - 420, h - 60))
        w, h = img.size

        ###################
        #     Damage      #
        ###################

        # We keep 75 to make sure name exists
        dmg_img = img.crop((300, 75, 600, h))
        cv2_dmg = cv2.cvtColor(np.array(dmg_img), cv2.COLOR_RGB2HSV)
        masked_dmg = cv2.inRange(cv2_dmg, (0, 0, 230), (255, 50, 255))
        raw_dmg = pytesseract.image_to_data(masked_dmg)
        stream = io.StringIO(raw_dmg)
        new_df = pd.read_csv(stream, sep='\t', quoting=csv.QUOTE_NONE)
        new_df = new_df[['left', 'top', 'conf', 'text']]
        new_df = new_df[new_df['text'].notnull()]
        new_df['text'] = new_df['text'].map(clean_dmg)
        if not len(new_df):
            continue

        dmg_left = dmg_left or new_df.iloc[0]['left']
        new_df = new_df[(new_df['conf'] > 80) &
                        (dmg_left - 5 < new_df['left']) &
                        (new_df['left'] < dmg_left + 5)]
        extra = new_df[(new_df['text'] < 20000) & (new_df['text'] >= 20)]

        # If the same number appears twice and is <20k, count it anyways
        extra_appeared = extra[extra['text'].isin(appeared)]
        appeared |= set(extra['text'].tolist())

        # If a damage appears twice, add it
        duplicates = new_df[new_df.duplicated(subset='text', keep=False)]

        new_df = new_df[new_df['text'] >= 20000]
        new_df = new_df.append(extra_appeared, ignore_index=True)

        new_df = new_df[~new_df['text'].isin(df['damage'])]
        new_df = new_df.append(duplicates, ignore_index=True)

        for _, entry in new_df.iterrows():

            ################
            #     Name     #
            ################

            name_img = img.crop((0, entry['top'], 300, entry['top'] + 50))
            cv2_name = cv2.cvtColor(np.array(name_img), cv2.COLOR_RGB2HSV)
            masked_name = cv2.inRange(cv2_name, (0, 0, 215), (255, 60, 255))
            raw_name = pytesseract.image_to_string(masked_name,
                                                   config="--psm 7")

            name = extract_name(raw_name)
            if not name:
                cv2.imwrite(f"ocr/error{error}.jpg", masked_name)
                error += 1

            ################
            #     Boss     #
            ################

            boss_img = img.crop((w - 450, entry['top'], w, entry['top'] + 50))
            cv2_boss = cv2.cvtColor(np.array(boss_img), cv2.COLOR_RGB2HSV)
            masked_boss = cv2.inRange(cv2_boss, (100, 0, 185), (180, 100, 255))
            raw_boss = pytesseract.image_to_string(masked_boss,
                                                   config="--psm 7")

            boss, level = extract_boss(raw_boss)
            if not boss:
                cv2.imwrite(f"ocr/error{error}.jpg", masked_boss)
                error += 1

            df = df.append({"name": name,
                            "damage": entry['text'],
                            "boss": boss,
                            "level": level},
                           ignore_index=True)

    df.drop_duplicates(subset=['name', 'damage', 'boss'], inplace=True)
    df = df.iloc[::-1]

    date_str = f"{dt.now():%m%d}"
    wb = openpyxl.load_workbook("data.xlsx")
    if date_str in wb.sheetnames:
        ws = wb[date_str]
        wb.remove(ws)

    ws = wb.create_sheet(date_str)

    for r in dataframe_to_rows(df, index=False):
        ws.append(r)

    for cell in ws[1]:
        cell.style = 'Pandas'

    error = iter(range(1, error))
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value is None:
                img = xlImage(f"ocr/error{next(error)}.jpg")
                img.anchor = cell.coordinate
                ws.add_image(img)

    wb.save("data.xlsx")
